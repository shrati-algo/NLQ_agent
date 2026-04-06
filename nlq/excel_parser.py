from openpyxl import load_workbook
from typing import Optional, Dict, Any, List


def parse_excel(path: str, sheet_name: Optional[str] = None, formula_cap: int = 200) -> Dict[str, Any]:
    """
    Parses Excel file and returns:
    - sheet name
    - headers
    - header_to_col mapping (Header -> Excel column letter)
    - col_to_header mapping (Excel column letter -> Header)
    - formula cells with metadata
    """

    wb = load_workbook(path, data_only=False)  # keep formulas
    ws = wb[sheet_name] if sheet_name else wb.active

    # ----------------------------
    # Read headers (Row 1)
    # ----------------------------
    headers: List[str] = []
    header_to_col: Dict[str, str] = {}
    col_to_header: Dict[str, str] = {}

    for cell in ws[1]:
        if cell.value is None:
            continue

        header = str(cell.value).strip()
        headers.append(header)

        col_letter = cell.column_letter
        header_to_col[header] = col_letter
        col_to_header[col_letter] = header

    # ----------------------------
    # Extract formulas
    # ----------------------------
    formulas = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                header = col_to_header.get(cell.column_letter)

                formulas.append({
                    "cell": cell.coordinate,            # e.g. H2
                    "header": header,                  # e.g. Width_Diff
                    "col_letter": cell.column_letter,  # e.g. H
                    "row": cell.row,                   # e.g. 2
                    "formula": cell.value              # e.g. =B2-C2
                })

                if len(formulas) >= formula_cap:
                    break
        if len(formulas) >= formula_cap:
            break

    return {
        "sheet": ws.title,
        "headers": headers,
        "header_to_col": header_to_col,
        "col_to_header": col_to_header,
        "formulas": formulas
    }
